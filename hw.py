# This is a sample Python script.
import statistics

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.
import cv2 as cv
import numpy as np
import os as os
from openpyxl import load_workbook as lwb
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import colorsys


#import excelFunc


def print_hi(name):
    # Use a breakpoint in the code line below to debug your script.
    print(f'Hi, {name}')  # Press Ctrl+F8 to toggle the breakpoint.


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    print_hi('PyCharm')


def getLBPMatrix(img,chanel):
    if chanel in range(0,3):
        result_matrix=img[:,:,chanel]*1.0
    else:
        result_matrix = 0.299 * img[:, :, 0] + 0.587 * img[:, :, 1] + 0.114 * img[:, :, 2]

    mean_delta=getMeanDelta(result_matrix)
    lbp = np.zeros((result_matrix.shape[0], result_matrix.shape[1]))
    for i in range(result_matrix.shape[0]):
       for j in range(result_matrix.shape[1]):
            # Calculate LBP for red channel
            data=((result_matrix[i, j] - result_matrix[max(0, i-1):min(result_matrix.shape[0], i+2), max(0, j-1):min(result_matrix.shape[1], j+2)]) <= mean_delta*1)
            c=np.zeros((3,3))
            s = data.size
            match s :
                case 4:
                    if i==0 and j==0:#Левый верхний угол
                       c[1:3, 1:3] = data
                    elif i!=0 and j!=0:#правый нижний угол
                       c[0:2, 0:2] = data
                    elif i!=0 and j==0:#Левый нижний угол
                        c[0:2, 1:3] = data
                    else:
                        c[1:3, 0:2] = data#правый верхний угол
                case 6 :
                    if i==result_matrix.shape[0]-1 and j>0:#нижняя сторона
                        c[0:2,0:3]=data
                    elif i==0 and j>0:#верхняя сторона
                        c[1:3, 0:4] = data
                    elif i!=0 and j==0:
                        c[0:4, 1:4] = data#левая сторона
                    elif i!=0 and j==result_matrix.shape[1]-1:
                        c[0:4, 0:2] = data  # правая сторона
                case _:
                    c[0:4,0:4]=data

            tmpV=np.zeros(8)
            jj=0
            for ii in range(0,c.shape[1]):
                tmpV[jj]=c[0,ii]
                jj=jj+1
            for ii in range(1, c.shape[0]):
                tmpV[jj]=c[ii,c.shape[0]-1]
                jj=jj+1
            for ii in range(c.shape[1]-2,-1,-1):
                tmpV[jj]=c[c.shape[0]-1,ii]
                jj=jj+1
            for ii in range(c.shape[0]-2,0,-1):
                tmpV[jj] = c[ii,0]
            mnojiteli = (128, 64, 32, 16, 8, 4, 2, 1)
            lbp[i, j] = np.sum(tmpV * mnojiteli)
    return lbp
#end getLBPmatrix



def getGrayLBPMatrix(img):
    # Convert RGB image to grayscale
    gray = 0.299*img[:,:,0] + 0.587*img[:,:,1] + 0.114*img[:,:,2]
    lbp = np.zeros((gray.shape[0], gray.shape[1]))
    for i in range(gray.shape[0]):
       for j in range(gray.shape[1]):
            # Calculate LBP for red channel
            data=((gray[i, j] - gray[max(0, i-1):min(gray.shape[0], i+2), max(0, j-1):min(gray.shape[1], j+2)]) >= 0)
            c=np.zeros((3,3))
            s = data.size
            match s :
                case 4:
                    if i==0 and j==0:#Левый верхний угол
                       c[1:3, 1:3] = data
                    elif i!=0 and j!=0:#правый нижний угол
                       c[0:2, 0:2] = data
                    elif i!=0 and j==0:#Левый нижний угол
                        c[0:2, 1:3] = data
                    else:
                        c[1:3, 0:2] = data#правый верхний угол
                case 6 :
                    if i==gray.shape[0]-1 and j>0:#нижняя сторона
                        c[0:2,0:3]=data
                    elif i==0 and j>0:#верхняя сторона
                        c[1:3, 0:4] = data
                    elif i!=0 and j==0:
                        c[0:4, 1:4] = data#левая сторона
                    elif i!=0 and j==gray.shape[1]-1:
                        c[0:4, 0:2] = data  # правая сторона
                case _:
                    c[0:4,0:4]=data

            tmpV=np.zeros(8)
            jj=0
            for ii in range(0,c.shape[1]):
                tmpV[jj]=c[0,ii]
                jj=jj+1
            for ii in range(1, c.shape[0]):
                tmpV[jj]=c[ii,c.shape[0]-1]
                jj=jj+1
            for ii in range(c.shape[1]-2,-1,-1):
                tmpV[jj]=c[c.shape[0]-1,ii]
                jj=jj+1
            for ii in range(c.shape[0]-2,0,-1):
                tmpV[jj] = c[ii,0]
            lbp[i, j] = np.sum(tmpV * 2 ** np.arange(8))
    return lbp
#end getGrayLBPmatrix


def getUniFormLBPHist(matr):
    uniform_lbp= [1, 2, 3, 4, 6, 7, 8, 12, 14, 15, 16, 24, 28, 30, 31, 32, 48, 56, 60, 62, 63, 64, 96, 112, 120,
                       124, 126, 127, 128, 129, 131, 135, 143, 159, 191, 192, 193, 195, 199, 207, 223, 224, 225, 227,
                       231, 239, 240, 241, 243, 247, 248, 249, 251, 252, 253, 254]
    bins = [1, 2, 3, 4, 6, 7, 8, 12, 14, 15, 16, 24, 28, 30, 31, 32, 48, 56, 60, 62, 63, 64, 96, 112, 120,
                       124, 126, 127, 128, 129, 131, 135, 143, 159, 191, 192, 193, 195, 199, 207, 223, 224, 225, 227,
                       231, 239, 240, 241, 243, 247, 248, 249, 251, 252, 253, 254,255,256]
    for i in range(0, matr.shape[0]):
        for j in range(0,matr.shape[1]):
            if matr[i, j] not in uniform_lbp:
                matr[i,j]=255
    lbpHist, bins = np.histogram(a=matr,bins=bins,density=False)
    #lbpHist=lbpHist*100
    return lbpHist, bins
#end     getUnoFormLBPHist

def copyPartMatrix(matr, leftX, rightX, topY, baseY):
    result = matr[topY:baseY+1, leftX:rightX+1, 0:3]
    return result
#end     copyPartMatrix

def getLBPPropsForPieceWithOverlap(matr,blok_size):
    # в зависимости от размера изображения расчитываем количестов итераций и размер вектора свойств для слоя
    end_w_iteration = (matr.shape[1] // blok_size)
    if end_w_iteration == 0:
        w_over = 0
        end_w_iteration = 1
    else:
        w_over = ((matr.shape[1]) - end_w_iteration * blok_size)
    end_w_iteration = end_w_iteration * 2 - 1
    end_h_iteration = ((matr.shape[0]) // blok_size)
    if end_h_iteration == 0:
        h_over = 0
        end_h_iteration = 1
    else:
        h_over = (matr.shape[0]) - end_h_iteration * blok_size
    end_h_iteration = end_h_iteration*2-1
    gray_lbp_list = []
    hsv_list = []
    #tmpLBPMatrix = np.zeros((int(end_h_iteration * blok_size), int(end_w_iteration * blok_size)))
    for ii in range(0, end_h_iteration):
        for jj in range(0, end_w_iteration):
            cyr_blok_matrix = copyPartMatrix(matr, jj * blok_size//2, jj * blok_size//2 + (blok_size - 1), ii * blok_size//2,
                                             ii * blok_size//2 + (blok_size - 1))
            hsv_props=get_HSV_props(cyr_blok_matrix)
            for k in range(0, len(hsv_props)):
                hsv_list.append(hsv_props[k])
            gray_blok_matrix = getLBPMatrix(cyr_blok_matrix, 3)
            gray_LBP_hist, gray_LBP_bin = getUniFormLBPHist(gray_blok_matrix)
            for k in range(0, gray_LBP_hist.shape[0]):
                gray_lbp_list.append(gray_LBP_hist[k])
        if w_over > 3:
            cyr_blok_matrix = copyPartMatrix(matr, matr.shape[1] - w_over + 1, matr.shape[1] - 1, ii * blok_size//2,
                                             ii * blok_size//2 + (blok_size - 1))
            hsv_props = get_HSV_props(cyr_blok_matrix)
            for k in range(0, len(hsv_props)):
                hsv_list.append(hsv_props[k])
            gray_blok_matrix = getLBPMatrix(cyr_blok_matrix, 3)
            gray_LBP_hist, gray_LBP_bin = getUniFormLBPHist(gray_blok_matrix)
            for k in range(0, gray_LBP_hist.shape[0]):
                gray_lbp_list.append(gray_LBP_hist[k])
    if h_over > 3:
        for jj in range(0, end_w_iteration):
            cyr_blok_matrix = copyPartMatrix(matr, jj * blok_size//2, jj * blok_size//2 + (blok_size - 1),
                                             matr.shape[0] - h_over + 1, matr.shape[0] - 1)
            hsv_props = get_HSV_props(cyr_blok_matrix)
            for k in range(0, len(hsv_props)):
                hsv_list.append(hsv_props[k])
            gray_blok_matrix = getLBPMatrix(cyr_blok_matrix, 3)
            gray_LBP_hist, gray_LBP_bin = getUniFormLBPHist(gray_blok_matrix)
            for k in range(0, gray_LBP_hist.shape[0]):
                gray_lbp_list.append(gray_LBP_hist[k])
    if w_over > 3:
        cyr_blok_matrix = copyPartMatrix(matr, matr.shape[1] - w_over + 1, matr.shape[1] - 1,
                                         matr.shape[0] - h_over + 1,
                                         matr.shape[0] - 1)
        hsv_props = get_HSV_props(cyr_blok_matrix)
        for k in range(0, len(hsv_props)):
            hsv_list.append(hsv_props[k])
        gray_blok_matrix = getLBPMatrix(cyr_blok_matrix, 3)
        gray_LBP_hist, gray_LBP_bin = getUniFormLBPHist(gray_blok_matrix)
        for k in range(0, gray_LBP_hist.shape[0]):
            gray_lbp_list.append(gray_LBP_hist[k])
    return gray_lbp_list, hsv_list

#end def getLBPPropsForPieceWithOverlap(matr, blok_size):

def getLBPPropsForPiece(matr, blok_size, tmpName):
    # в зависимости от размера изображения расчитываем количестов итераций и размер вектора свойств для слоя
    end_w_iteration = matr.shape[1] // blok_size
    if end_w_iteration == 0:
        w_over = 0
        end_w_iteration = 1
    else:
        w_over = ((matr.shape[1]) - end_w_iteration * blok_size)
    end_h_iteration = (matr.shape[0]) // blok_size
    if end_h_iteration == 0:
        h_over = 0
        end_h_iteration = 1
    else:
        h_over = (matr.shape[0]) - end_h_iteration * blok_size
    '''
    if w_over > 3:
        if h_over > 3:
            gray_props_vector = np.zeros(((end_w_iteration + 1) * (end_h_iteration + 1)) * 57)
            red_props_vector = np.zeros(((end_w_iteration + 1) * (end_h_iteration + 1)) * 57)
            green_props_vector = np.zeros(((end_w_iteration + 1) * (end_h_iteration + 1)) * 57)
            blue_props_vector = np.zeros(((end_w_iteration + 1) * (end_h_iteration + 1)) * 57)
        else:
            gray_props_vector = np.zeros((end_w_iteration + 1) * end_h_iteration * 57)
            red_props_vector = np.zeros((end_w_iteration + 1) * end_h_iteration * 57)
            green_props_vector = np.zeros((end_w_iteration + 1) * end_h_iteration * 57)
            blue_props_vector = np.zeros((end_w_iteration + 1) * end_h_iteration * 57)
    else:
        if h_over > 3:
            gray_props_vector = np.zeros(end_w_iteration * (end_h_iteration + 1) * 57)
            red_props_vector = np.zeros(end_w_iteration * (end_h_iteration + 1) * 57)
            green_props_vector = np.zeros(end_w_iteration * (end_h_iteration + 1) * 57)
            blue_props_vector = np.zeros(end_w_iteration * (end_h_iteration + 1) * 57)
        else:
            gray_props_vector = np.zeros(end_w_iteration * end_h_iteration * 57)
            red_props_vector = np.zeros(end_w_iteration * end_h_iteration * 57)
            green_props_vector = np.zeros(end_w_iteration * end_h_iteration * 57)
            blue_props_vector = np.zeros(end_w_iteration * end_h_iteration * 57)
    '''
    gray_lbp_list = []
    red_lbp_list = []
    green_lbp_list = []
    blue_lbp_list = []
    tmpLBPMatrix = np.zeros((int(end_h_iteration*blok_size), int(end_w_iteration*blok_size)))
    for ii in range(0, end_h_iteration):
        for jj in range(0, end_w_iteration):
            cyr_blok_matrix = copyPartMatrix(matr, jj * blok_size, jj * blok_size + (blok_size - 1), ii * blok_size,
                                           ii * blok_size + (blok_size - 1))
            gray_blok_matrix = getLBPMatrix(cyr_blok_matrix,3)
            tmpLBPMatrix[ii*gray_blok_matrix.shape[0]:ii*gray_blok_matrix.shape[0]+gray_blok_matrix.shape[0],jj*gray_blok_matrix.shape[1]:jj*gray_blok_matrix.shape[1]+gray_blok_matrix.shape[1]]=gray_blok_matrix
            # red_blok_matrix = getLBPMatrix(cyr_blok_matrix, 0)
            # green_blok_matrix = getLBPMatrix(cyr_blok_matrix, 1)
            # blue_blok_matrix = getLBPMatrix(cyr_blok_matrix, 2)
            gray_LBP_hist, gray_LBP_bin = getUniFormLBPHist(gray_blok_matrix)
            # red_LBP_hist, red_LBP_bin = getUniFormLBPHist(red_blok_matrix)
            # green_LBP_hist, green_LBP_bin = getUniFormLBPHist(green_blok_matrix)
            # blue_LBP_hist, blue_LBP_bin = getUniFormLBPHist(blue_blok_matrix)
            for k in range(0, gray_LBP_hist.shape[0]):
                gray_lbp_list.append(gray_LBP_hist[k])
                # red_lbp_list.append(red_LBP_hist[k])
                # green_lbp_list.append(green_LBP_hist[k])
                # blue_lbp_list.append(blue_LBP_hist[k])
        if w_over > 3:
            cyr_blok_matrix = copyPartMatrix(matr, matr.shape[1] - w_over + 1, matr.shape[1] - 1, ii * blok_size,
                                           ii * blok_size + (blok_size - 1))
            gray_blok_matrix = getLBPMatrix(cyr_blok_matrix,3)
            # red_blok_matrix = getLBPMatrix(cyr_blok_matrix, 0)
            # green_blok_matrix = getLBPMatrix(cyr_blok_matrix, 1)
            # blue_blok_matrix = getLBPMatrix(cyr_blok_matrix, 2)
            gray_LBP_hist, gray_LBP_bin = getUniFormLBPHist(gray_blok_matrix)
            # red_LBP_hist, red_LBP_bin = getUniFormLBPHist(red_blok_matrix)
            # green_LBP_hist, green_LBP_bin = getUniFormLBPHist(green_blok_matrix)
            # blue_LBP_hist, blue_LBP_bin = getUniFormLBPHist(blue_blok_matrix)
            for k in range(0, gray_LBP_hist.shape[0]):
                gray_lbp_list.append(gray_LBP_hist[k])
                # red_lbp_list.append(red_LBP_hist[k])
                # green_lbp_list.append(green_LBP_hist[k])
                # blue_lbp_list.append(blue_LBP_hist[k])
    if h_over > 3:
        for jj in range(0, end_w_iteration):
            cyr_blok_matrix = copyPartMatrix(matr, jj * blok_size, jj * blok_size + (blok_size - 1),
                                           matr.shape[0] - h_over + 1, matr.shape[0] - 1)
            gray_blok_matrix = getLBPMatrix(cyr_blok_matrix,3)
            # red_blok_matrix = getLBPMatrix(cyr_blok_matrix, 0)
            # green_blok_matrix = getLBPMatrix(cyr_blok_matrix, 1)
            # blue_blok_matrix = getLBPMatrix(cyr_blok_matrix, 2)
            gray_LBP_hist, gray_LBP_bin = getUniFormLBPHist(gray_blok_matrix)
            # red_LBP_hist, red_LBP_bin = getUniFormLBPHist(red_blok_matrix)
            # green_LBP_hist, green_LBP_bin = getUniFormLBPHist(green_blok_matrix)
            # blue_LBP_hist, blue_LBP_bin = getUniFormLBPHist(blue_blok_matrix)
            for k in range(0, gray_LBP_hist.shape[0]):
                gray_lbp_list.append(gray_LBP_hist[k])
                # red_lbp_list.append(red_LBP_hist[k])
                # green_lbp_list.append(green_LBP_hist[k])
                # blue_lbp_list.append(blue_LBP_hist[k])
    if w_over > 3:
        cyr_blok_matrix = copyPartMatrix(matr, matr.shape[1] - w_over + 1, matr.shape[1] - 1, matr.shape[0] - h_over + 1,
                                       matr.shape[0] - 1)
        gray_blok_matrix = getLBPMatrix(cyr_blok_matrix,3)
        # red_blok_matrix = getLBPMatrix(cyr_blok_matrix, 0)
        # green_blok_matrix = getLBPMatrix(cyr_blok_matrix, 1)
        # blue_blok_matrix = getLBPMatrix(cyr_blok_matrix, 2)
        gray_LBP_hist, gray_LBP_bin = getUniFormLBPHist(gray_blok_matrix)
        # red_LBP_hist, red_LBP_bin = getUniFormLBPHist(red_blok_matrix)
        # green_LBP_hist, green_LBP_bin = getUniFormLBPHist(green_blok_matrix)
        # blue_LBP_hist, blue_LBP_bin = getUniFormLBPHist(blue_blok_matrix)
        for k in range(0, gray_LBP_hist.shape[0]):
            gray_lbp_list.append(gray_LBP_hist[k])
            # red_lbp_list.append(red_LBP_hist[k])
            # green_lbp_list.append(green_LBP_hist[k])
            # blue_lbp_list.append(blue_LBP_hist[k])
    cv.imwrite("./tmp/" + tmpName + "_lbp.jpg", tmpLBPMatrix)
    return gray_lbp_list, red_lbp_list, green_lbp_list, blue_lbp_list

#end def getLBPPropsForPiece

def getLBPPropsForLayer(matr, blokSize):
# в зависимости от размера изображения расчитываем количестов итераций и размер вектора свойств для слоя
    endWIteration = matr.shape[1]//blokSize
    if endWIteration == 0:
        wOver = 0
        endWIteration=1
    else:
        wOver = ((matr.shape[1] ) - endWIteration * blokSize)
    endHIteration = (matr.shape[0])//blokSize
    if endHIteration == 0:
        hOver = 0
        endHIteration=1
    else:
        hOver = (matr.shape[0] ) - endHIteration * blokSize
    if wOver > 3:
        if hOver > 3:
                cyrPiecePropVector = np.zeros(((endWIteration + 1) * (endHIteration + 1))*57)
        else:
                cyrPiecePropVector = np.zeros((endWIteration + 1) * endHIteration * 57)
    else:
        if hOver > 3:
            cyrPiecePropVector = np.zeros(endWIteration * (endHIteration+1) * 57)
        else:
            cyrPiecePropVector = np.zeros(endWIteration * endHIteration * 57)

    lbpList=[]
    for ii in range(0,endHIteration):
        for jj in range(0,endWIteration):
            cyrBlokMatrix = copyPartMatrix(matr,jj * blokSize, jj * blokSize + (blokSize - 1), ii * blokSize, ii * blokSize + (blokSize - 1))
            cyrBlokMatrix=getGrayLBPMatrix(cyrBlokMatrix)
            cyrLBPHist, cyrLBPBin=getUniFormLBPHist(cyrBlokMatrix)
            #cyrPiecePropVector[(ii+jj) * 57:(ii+jj) * 57 + 57]=cyrLBPHist
            for k in range(0,cyrLBPHist.shape[0]):
                lbpList.append(cyrLBPHist[k])
        if wOver > 3:
            cyrBlokMatrix = copyPartMatrix(matr, matr.shape[1] - wOver + 1, matr.shape[1] - 1, ii * blokSize, ii * blokSize + (blokSize - 1))
            cyrBlokMatrix = getGrayLBPMatrix(cyrBlokMatrix)
            cyrLBPHist, cyrLBPBin = getUniFormLBPHist(cyrBlokMatrix)
            #cyrPiecePropVector[(ii+jj+1) * 57:(ii+jj+1) * 57 + 57] = cyrLBPHist
            for k in range(0,cyrLBPHist.shape[0]):
                lbpList.append(cyrLBPHist[k])
    if hOver > 3:
        for jj in range(0, endWIteration):
            cyrBlokMatrix = copyPartMatrix(matr, jj * blokSize, jj * blokSize + (blokSize - 1), matr.shape[0]-hOver+1, matr.shape[0]-1)
            cyrBlokMatrix = getGrayLBPMatrix(cyrBlokMatrix)
            cyrLBPHist, cyrLBPBin = getUniFormLBPHist(cyrBlokMatrix)
            #cyrPiecePropVector[(ii+jj) * 57:(ii+jj) * 57 + 57] = cyrLBPHist
            for k in range(0,cyrLBPHist.shape[0]):
                lbpList.append(cyrLBPHist[k])
    if wOver > 3:
        cyrBlokMatrix = copyPartMatrix(matr, matr.shape[1]-wOver+1, matr.shape[1]-1, matr.shape[0] - hOver + 1, matr.shape[0] - 1)
        cyrBlokMatrix = getGrayLBPMatrix(cyrBlokMatrix)
        cyrLBPHist, cyrLBPBin = getUniFormLBPHist(cyrBlokMatrix)
        #cyrPiecePropVector[(ii+jj) * 57:(ii+jj) * 57 + 57] = cyrLBPHist
        for k in range(0, cyrLBPHist.shape[0]):
            lbpList.append(cyrLBPHist[k])
    return lbpList
#end getPropsForLayer(matr)

def getCOSSimilarity(v1,v2):
    result = np.dot(v1,v2)/(np.linalg.norm(v1) * np.linalg.norm(v2))
    return result
#end getCOSSimilarity

def getMeanDelta(matr):
    mean_delta=0.0
    for i in range(1, matr.shape[0] -1):
        for j in range(1, matr.shape[1]-1):
            mean_delta =mean_delta+ (abs(matr[i, j] - matr[i, j - 1]) + abs(matr[i, j] - matr[i - 1, j - 1]) +
                           abs(matr[i, j] - matr[i - 1, j]) + abs(matr[i, j] - matr[i - 1, j + 1])) / 4.0
    mean_delta /= (matr.shape[1]-2 ) * (matr.shape[0]-2 )  # normalizing
    mean_delta *=  -1
    return mean_delta
#end def getMeanDelta(matr):

def distance_between_vectors(A, B):
    distance = np.sqrt(np.sum((B - A)**2))
    return distance
#end distance_between_vectors()

def angle_between_vectors(A, B):
    dot_product = np.sum(A * B)
    magnitude_A = np.linalg.norm(A)
    magnitude_B = np.linalg.norm(B)
    cos_theta = dot_product / (magnitude_A * magnitude_B)
    theta = np.arccos(cos_theta)
    return theta

def get_HSV_props(matr):
    h_comp = []
    s_comp = []
    v_comp = []
    for i in range(matr.shape[0]):
        for j in range(matr.shape[1]):
            hsv =colorsys.rgb_to_hsv(matr[i,j,0],matr[i,j,1],matr[i,j,2])
            h_comp.append(hsv[0])
            s_comp.append(hsv[1])
            v_comp.append(hsv[2])
    h_comp=np.array(h_comp)
    s_comp = np.array(s_comp)
    v_comp = np.array(v_comp)
    HSV_props=[]
    HSV_props.append(h_comp.mean())
    HSV_props.append(s_comp.mean())
    HSV_props.append(v_comp.mean())
    HSV_props.append(h_comp.std())
    HSV_props.append(s_comp.std())
    HSV_props.append(v_comp.std())
    HSV_props.append(h_comp.var())
    HSV_props.append(s_comp.var())
    HSV_props.append(v_comp.var())
    HSV_props.append(statistics.mode(h_comp))
    HSV_props.append(statistics.mode(s_comp))
    HSV_props.append(statistics.mode(v_comp))
    HSV_props.append(statistics.median(h_comp))
    HSV_props.append(statistics.median(s_comp))
    HSV_props.append(statistics.median(v_comp))
    return HSV_props

def get_RGB_props(matr):
    r_comp = []
    g_comp = []
    b_comp = []
    for i in range(matr.shape[0]):
        for j in range(matr.shape[1]):
            rgb =(matr[i,j,0],matr[i,j,1],matr[i,j,2])
            r_comp.append(rgb[0])
            g_comp.append(rgb[1])
            b_comp.append(rgb[2])
    r_comp=np.array(r_comp)
    g_comp = np.array(g_comp)
    b_comp = np.array(b_comp)
    RGB_props=[]
    RGB_props.append(r_comp.mean())
    RGB_props.append(g_comp.mean())
    RGB_props.append(b_comp.mean())
    # HSV_props.append(h_comp.std())
    # HSV_props.append(s_comp.std())
    # HSV_props.append(v_comp.std())
    # HSV_props.append(h_comp.var())
    # HSV_props.append(s_comp.var())
    # HSV_props.append(v_comp.var())
    # HSV_props.append(statistics.mode(h_comp))
    # HSV_props.append(statistics.mode(s_comp))
    # HSV_props.append(statistics.mode(v_comp))
    # HSV_props.append(statistics.median(h_comp))
    # HSV_props.append(statistics.median(s_comp))
    # HSV_props.append(statistics.median(v_comp))
    return RGB_props




wb= load_workbook('./Book1.xlsx')
sheet = wb.worksheets[0]
print(type(sheet))
gray_prop_vect_arr =[]
hsv_prop_vect_arr=[]
red_prop_vect_arr =[]
green_prop_vect_arr =[]
blue_prop_vect_arr =[]
blok_size=32

start_row=3
end_row=18
for cellObj in sheet['A'+str(start_row):'A'+str(end_row)]:
      for cell in cellObj:
              image_C=cv.imread("C:/Users/user/Documents/GitHub/myRepo2/FotoCore/"+cell.value)
              top_Y=sheet.cell(cell.row,13).value
              base_Y = sheet.cell(cell.row, 14).value
              image_C=copyPartMatrix(image_C, 10, 138, top_Y, base_Y)
              cv.imwrite("./tmp/"+cell.value[0:len(cell.value)-4]+".jpg", image_C)
              gray_prop_V, HSV_prop_v = getLBPPropsForPieceWithOverlap(image_C, blok_size)
              gray_prop_vect_arr.append(gray_prop_V)
              hsv_prop_vect_arr.append(HSV_prop_v)
              print('property vectors '+cell.value, gray_prop_V)
              print('property vectors '+cell.value, HSV_prop_v)
#-----------------------------------------------------------------------------------
#добиваем нулями
for_cluster=[]
cos_similarity_arr = []
for i in range(0, len(gray_prop_vect_arr)-1):
    gray_prop_v1=gray_prop_vect_arr[i]
    gray_prop_v2=gray_prop_vect_arr[i+1]
    set1 = set(gray_prop_v1)
    set2 = set(gray_prop_v2)
    jacard_similarity = len(set1 & set2 )/len(set1 | set2)


    if len(gray_prop_v1)>=len(gray_prop_v2):
        if len(gray_prop_v1)/len(gray_prop_v2)<2:
            tmp_prop_v=np.zeros(len(gray_prop_v1))
            tmp_prop_v[0:len(gray_prop_v2)]=gray_prop_v2[0:len(gray_prop_v2)]
            cos_similarity = getCOSSimilarity(gray_prop_v1, tmp_prop_v)
            dist = distance_between_vectors(gray_prop_v1, tmp_prop_v)
            angle =  angle_between_vectors(gray_prop_v1, tmp_prop_v)
        else:
            end_h=len(gray_prop_v1)//len(gray_prop_v2)
            cos_similarity=-1
            for ii in range(0,1):
                tmp_prop_v = np.zeros(len(gray_prop_v2))
                tmp_prop_v[0:len(gray_prop_v2)] = gray_prop_v1[0:len(gray_prop_v2)]
                tmp_cos_similarity = getCOSSimilarity(gray_prop_v2, tmp_prop_v)
                dist = distance_between_vectors(gray_prop_v2, tmp_prop_v)
                angle = angle_between_vectors(gray_prop_v2, tmp_prop_v)
                if tmp_cos_similarity>cos_similarity:
                    cos_similarity=tmp_cos_similarity
    else:
        if len(gray_prop_v2) / len(gray_prop_v1) < 2:
            tmp_prop_v=np.zeros(len(gray_prop_v2))
            tmp_prop_v[0:len(gray_prop_v1)]=gray_prop_v1[0:len(gray_prop_v1)]
            cos_similarity = getCOSSimilarity(gray_prop_v2, tmp_prop_v)
            dist = distance_between_vectors(gray_prop_v2, tmp_prop_v)
            angle = angle_between_vectors(gray_prop_v2, tmp_prop_v)
        else:
            end_h=len(gray_prop_v2)//len(gray_prop_v1)
            cos_similarity=-1
            for ii in range(0,1):
                tmp_prop_v = np.zeros(len(gray_prop_v1))
                tmp_prop_v[0:len(gray_prop_v1)] = gray_prop_v2[0:len(gray_prop_v1)]
                tmp_cos_similarity = getCOSSimilarity(gray_prop_v1, tmp_prop_v)
                dist = distance_between_vectors(gray_prop_v1, tmp_prop_v)
                angle = angle_between_vectors(gray_prop_v1, tmp_prop_v)
                if tmp_cos_similarity>cos_similarity:
                    cos_similarity=tmp_cos_similarity
    print(sheet.cell(i+3,1).value,"-",sheet.cell(i+4,1).value,"   ",round(cos_similarity,5), "  ", dist, "    ", angle/(np.pi/180), "jacard-  ",(cos_similarity+jacard_similarity)/2 )
    for_cluster.append([dist,angle])
    cos_similarity_arr.append([sheet.cell(i+start_row,1).value + "-" +sheet.cell(i+start_row+1,1).value, cos_similarity])
    #if round(cos_similarity,1) < 0.9:
    #print(sheet.cell(i+start_row,1).value,"-",sheet.cell(i+start_row+1,1).value,"                ",cos_similarity)

X=np.array(for_cluster)
y_preds=clusterKraftTest.run_Kmeans(2,X)
for i in range(0,len(y_preds)):
    if round(cos_similarity_arr[i][1],2) < 0.95:
        cos_similarity_arr[i][1]=cos_similarity_arr[i][1]*y_preds[i]

for cos_similarity_item in  cos_similarity_arr:
    if round(cos_similarity_item[1],2) < 0.95 :
        print(cos_similarity_item[0],"   ",cos_similarity_item[1])

for i in range(0, len(y_preds)):
    print(str(X[i])+" - "+ str(y_preds[i]))

#---------------------------------------------------------
#создаем спмсок масиввов для кластеризации
X_for_cluster = []
for i in range(0, len(gray_prop_vect_arr)):
    if len(gray_prop_vect_arr[i])<max_length:
        tmp_prop_v=np.zeros(max_length)
        tmp_prop_v[0:len(gray_prop_vect_arr[i])]=gray_prop_vect_arr[i]
        X_for_cluster.append(tmp_prop_v)
    else:
        X_for_cluster.append(gray_prop_vect_arr[i])


'''
#-----------------------------------------------------------------------------------
#конец добиваем нулями
print("")
#обрезаем длинный
for i in range(0, len(prop_vect_arr)-1):
    prop_v1=prop_vect_arr[i]
    prop_v2=prop_vect_arr[i+1]
    if len(prop_v1)>=len(prop_v2):
        tmp_prop_v=np.zeros(len(prop_v2))
        tmp_prop_v[0:len(prop_v2)]=prop_v1[0:len(prop_v2)]
        cos_similarity = getCOSSimilarity(prop_v2, tmp_prop_v)
    else:
        tmp_prop_v=np.zeros(len(prop_v1))
        tmp_prop_v[0:len(prop_v1)]=prop_v2[0:len(prop_v1)]
        cos_similarity = getCOSSimilarity(prop_v1, tmp_prop_v)
    #print(cos_similarity)
    if cos_similarity < 0.90:
       print(sheet.cell(i+65,1).value,"-",sheet.cell(i+66,1).value,"                ",cos_similarity)
#-----------------------------------------------------------------------------------
#конец обрезаем длинный
'''



# See PyCharm help at https://www.jetbrains.com/help/pycharm/
